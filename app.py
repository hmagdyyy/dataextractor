import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill as XLFill
from io import BytesIO
import re
from datetime import datetime

st.set_page_config(page_title="Mini Client Dashboard", layout="wide")
st.title("Mini Client Dashboard")

# -------------------------------------------------
# Helpers
# -------------------------------------------------
def normalize_stock(s: str) -> str:
    if not s:
        return ""
    s = str(s).strip().upper()
    return s if s.endswith(".CA") else f"{s}.CA"

def norm_client_name(s: str) -> str:
    return str(s).strip().title() if s is not None else ""

def export_xlsx(df: pd.DataFrame, filename="export.xlsx", sheet_name="Sheet1"):
    """Export a DataFrame as a single-sheet XLSX with simple numeric formatting."""
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        for col in ws.columns:
            for cell in col[1:]:
                if isinstance(cell.value, (int, float)):
                    if float(cell.value).is_integer():
                        cell.number_format = "#,##0"
                    else:
                        cell.number_format = "#,##0.00"
    buffer.seek(0)
    st.download_button(
        "📥 Download Excel",
        buffer,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

def parse_mixed_date(x):
    """
    If contains '-' => dd-mm-yyyy
    If contains '/' => mm/dd/yyyy
    Otherwise let pandas try.
    """
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return pd.NaT
    if isinstance(x, (datetime, pd.Timestamp)):
        return pd.to_datetime(x, errors="coerce")

    s = str(x).strip()
    if not s:
        return pd.NaT

    try:
        if "-" in s:
            return pd.to_datetime(s, errors="coerce", dayfirst=True)
        if "/" in s:
            return pd.to_datetime(s, errors="coerce", dayfirst=False)
        return pd.to_datetime(s, errors="coerce")
    except Exception:
        return pd.NaT

# -------------------------------------------------
# Cashflow (.xls) parser
# -------------------------------------------------
def load_cashflows_xls(xls_file):
    """
    Matches:
      - Entry Date: column A
      - Client: column D (title 'Client')
      - Type: column G (under Trans.Date)  In=deposit, Out=withdrawal
      - Amount: column I (under Status)
    Returns: dict client -> DataFrame[Entry Date, Type, Amount] sorted desc
    """
    if xls_file is None:
        return {}

    # read .xls robustly (requires xlrd)
    try:
        df = pd.read_excel(xls_file, sheet_name=0, engine="xlrd")
    except Exception:
        df = pd.read_excel(xls_file, sheet_name=0)

    if df is None or df.empty:
        return {}

    # if columns are too few, try header=None
    try:
        if df.shape[1] < 9:
            df = pd.read_excel(xls_file, sheet_name=0, header=None, engine="xlrd")
    except Exception:
        pass

    if df.shape[1] < 9:
        return {}

    entry_col = 0   # A
    client_col = 3  # D
    type_col = 6    # G
    amt_col = 8     # I

    out = {}
    for _, r in df.iterrows():
        client = r.iloc[client_col] if client_col < len(r) else None
        if client is None or str(client).strip() == "" or str(client).strip().lower() == "client":
            continue

        entry = r.iloc[entry_col] if entry_col < len(r) else None
        tx_type = r.iloc[type_col] if type_col < len(r) else None
        amt = r.iloc[amt_col] if amt_col < len(r) else None

        client_n = norm_client_name(client)
        dt = parse_mixed_date(entry)

        t_raw = str(tx_type).strip().lower() if tx_type is not None else ""
        if t_raw == "in":
            t = "In"
        elif t_raw == "out":
            t = "Out"
        else:
            t = str(tx_type).strip().title() if tx_type is not None else ""

        try:
            amt_f = float(amt)
        except Exception:
            amt_f = 0.0

        out.setdefault(client_n, []).append({"Entry Date": dt, "Type": t, "Amount": amt_f})

    dfs = {}
    for c, rows in out.items():
        cdf = pd.DataFrame(rows)
        if not cdf.empty:
            cdf["Entry Date"] = pd.to_datetime(cdf["Entry Date"], errors="coerce")
            cdf = cdf.sort_values("Entry Date", ascending=False).reset_index(drop=True)
        dfs[c] = cdf
    return dfs

def style_cashflows(df: pd.DataFrame):
    if df is None or df.empty:
        return df

    def row_style(row):
        t = str(row.get("Type", "")).strip().lower()
        if t == "in":
            return ["background-color: #d9f2d9"] * len(row)  # light green
        if t == "out":
            return ["background-color: #f7d6d6"] * len(row)  # light red
        return [""] * len(row)

    sdf = df.copy()
    return sdf.style.apply(row_style, axis=1).format({"Amount": "{:,.2f}"})


# -------------------------------------------------
# Extraction (main workbook)
# -------------------------------------------------
def extract_client_data(file):
    """
    Extracts:
      - Client: B4
      - Cash: C27
      - Dividends: C32
      - Fees Under Payment: Cxx where Bxx = 'Fees Under Payment'
      - Stocks: B=Name, C=Qty, D=Cost, E=Price, H=MV, I=Weight
      - ICs: scan next <=10 rows after stocks end for Stk-300 (Stream MV) & Stk-302 (Momentum MV)
      - AUM: 'Total Assets' row (col C)
      - Total Cash (legacy): Cash + Dividends + Stream(MV)
      - Prices table
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    out = {}
    all_prices = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        client = (ws["B4"].value or sheet_name).title()

        cash = float(ws["C27"].value or 0)
        dividends = float(ws["C32"].value or 0)

        # Fees Under Payment (scan rows 35..70, label in B, value in C)
        fees_under_payment = 0.0
        for r in ws.iter_rows(min_row=35, max_row=70):
            label = r[1].value
            if isinstance(label, str) and label.strip().lower() == "fees under payment":
                fees_under_payment = float(r[2].value or 0.0)
                break

        # Locate stocks
        start_row = None
        for r in ws.iter_rows(min_row=1, max_row=100):
            if str(r[0].value).strip().lower() == "stocks":
                start_row = r[0].row + 1
                break

        stock_rows = []
        stream_mv = 0.0
        momentum_mv = 0.0
        arope_mv = 0.0
        last_stock_row = None

        if start_row:
            for r in ws.iter_rows(min_row=start_row):
                cell_b = r[1]
                rgb = getattr(getattr(cell_b.fill, "start_color", None), "rgb", None)
                empty_b = (cell_b.value is None) or (str(cell_b.value).strip() == "")
                if rgb == "FFD3D3D3" and empty_b:
                    last_stock_row = r[0].row
                    break

                name = r[1].value
                qty = r[2].value
                cost = r[3].value if len(r) > 3 else 0
                price = r[4].value if len(r) > 4 else 0
                mv = r[7].value if len(r) > 7 else 0
                wt = r[8].value if len(r) > 8 else 0

                if not name:
                    continue

                name_str = str(name).strip()
                name_up = name_str.upper()

                if isinstance(price, (int, float)):
                    all_prices[normalize_stock(name_str)] = float(price)

                # ignore IC tickers here
                if name_up not in ("STK-300", "STK-302","STK-291") and isinstance(qty, (int, float)):
                    stock_rows.append({
                        "Company Name": normalize_stock(name_str),
                        "Quantity": float(qty or 0),
                        "Cost": float(cost or 0),
                        "Price": float(price or 0),
                        "Market Value": float(mv or 0),
                        "Weight": wt or 0,
                    })

            # Scan <=10 rows after stocks for ICs
            if not last_stock_row:
                last_stock_row = start_row
            ic_start, ic_end = last_stock_row + 1, last_stock_row + 11
            for r in ws.iter_rows(min_row=ic_start, max_row=ic_end):
                name_cell = r[1].value
                if not name_cell:
                    continue
                nm = str(name_cell).strip().upper()
                mv = r[7].value if len(r) > 7 else 0
                if nm == "STK-300":
                    stream_mv = float(mv or 0)
                elif nm == "STK-291":
                    arope_mv = float(mv or 0)
                elif nm == "STK-302":
                    momentum_mv = float(mv or 0)

        # AUM
        aum = 0.0
        for r in ws.iter_rows(min_row=1, max_row=100):
            if str(r[0].value).strip().lower() == "total assets":
                aum = float(r[2].value or 0)
                break

        total_cash = cash + dividends + stream_mv + arope_mv

        df_stocks = pd.DataFrame(
            stock_rows,
            columns=["Company Name", "Quantity", "Cost", "Price", "Market Value", "Weight"]
        )

        out[client] = {
            "data": df_stocks,
            "cash": cash,
            "dividends": dividends,
            "stream_mv": stream_mv,
            "arope_mv": arope_mv,
            "momentum_mv": momentum_mv,
            "total_cash": total_cash,
            "aum": aum,
            "fees_under_payment": fees_under_payment,
        }

    prices_df = (
        pd.DataFrame(sorted(all_prices.items()), columns=["Stock", "Price"])
        if all_prices else pd.DataFrame(columns=["Stock", "Price"])
    )
    return out, prices_df


# -------------------------------------------------
# Views
# -------------------------------------------------
def client_view(data, cashflows_by_client=None):
    client = st.selectbox("Select Client", sorted(data.keys()))
    info = data[client]

    st.subheader(client)

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Cash (C27)", f"{info['cash']:,.2f}")
    c2.metric("Dividends (C32)", f"{info['dividends']:,.2f}")
    c3.metric("Stream (Stk-300)", f"{info['stream_mv']:,.2f}")
    c4.metric("Momentum (Stk-302)", f"{info['momentum_mv']:,.2f}")
    c5.metric("Total Cash", f"{info['total_cash']:,.2f}")

    st.metric("AUM", f"{info['aum']:,.2f}")

    st.markdown("**Stock Holdings**")
    st.dataframe(info["data"], use_container_width=True, hide_index=True)
    export_xlsx(info["data"], filename=f"{client}_holdings.xlsx", sheet_name="Holdings")

    st.markdown("---")
    st.subheader("Cashflows (optional)")

    if cashflows_by_client and client in cashflows_by_client:
        tx = cashflows_by_client[client].copy()
        if tx is None or tx.empty:
            st.info("No cashflow transactions found for this client.")
        else:
            tx["Entry Date"] = pd.to_datetime(tx["Entry Date"], errors="coerce")
            tx = tx.sort_values("Entry Date", ascending=False).reset_index(drop=True)
            st.dataframe(style_cashflows(tx), use_container_width=True, hide_index=True)
            export_xlsx(tx, filename=f"{client}_cashflows.xlsx", sheet_name="Cashflows")
    else:
        st.info("Upload the Cashflow .xls file above to see transactions here (if available).")


def total_portfolio_view(data):
    st.subheader("Total Portfolio View")

    all_stocks = sorted({s for v in data.values() for s in v["data"]["Company Name"].unique()})
    rows = []

    for client, info in sorted(data.items()):
        raw_cash = float(info["cash"])
        fees = float(info.get("fees_under_payment", 0.0) or 0.0)
        net_cash = raw_cash - fees

        total_cash_view = net_cash + float(info["dividends"]) + float(info["stream_mv"]) + float(info["arope_mv"])

        row = {
            "Client": client,
            "Cash": net_cash,
            "Fees Under Payment": fees,
            "Dividends": info["dividends"],
            "Stream": info["stream_mv"],
            "Momentum": info["momentum_mv"],
            "Total Cash": total_cash_view,
            "NAV": info["aum"],
        }

        for s in all_stocks:
            row[f"{s} Qty"] = 0
            row[f"{s} Cost"] = 0

        for _, r in info["data"].iterrows():
            sym = r["Company Name"]
            row[f"{sym} Qty"] = float(r.get("Quantity", 0) or 0)
            row[f"{sym} Cost"] = float(r.get("Cost", 0) or 0)

        rows.append(row)

    fixed = ["Client", "Cash", "Fees Under Payment", "Dividends", "Stream", "Momentum", "Total Cash"]
    stock_cols = []
    for s in all_stocks:
        stock_cols += [f"{s} Qty", f"{s} Cost"]
    cols = fixed + stock_cols + ["NAV"]

    mat = pd.DataFrame(rows, columns=cols)
    st.dataframe(mat, use_container_width=True)
    export_xlsx(mat, filename="total_portfolio.xlsx", sheet_name="Portfolio")


def total_portfolio_view_weights(data):
    st.subheader("Total Portfolio View (Weights)")

    all_stocks = sorted({s for v in data.values() for s in v["data"]["Company Name"].unique()})
    rows = []

    for client, info in sorted(data.items()):
        raw_cash = float(info["cash"])
        fees = float(info.get("fees_under_payment", 0.0) or 0.0)
        net_cash = raw_cash - fees

        row = {
            "Client": client,
            "Cash": net_cash,
            "Fees Under Payment": fees,
            "NAV": info["aum"],
        }

        for s in all_stocks:
            row[f"{s} Wt"] = 0
            row[f"{s} Cost"] = 0

        for _, r in info["data"].iterrows():
            sym = r["Company Name"]
            row[f"{sym} Wt"] = float(r.get("Weight", 0) or 0)
            row[f"{sym} Cost"] = float(r.get("Cost", 0) or 0)

        rows.append(row)

    fixed = ["Client", "Cash", "Fees Under Payment"]
    stock_cols = []
    for s in all_stocks:
        stock_cols += [f"{s} Wt", f"{s} Cost"]
    cols = fixed + stock_cols + ["NAV"]

    mat = pd.DataFrame(rows, columns=cols)
    st.dataframe(mat, use_container_width=True)
    export_xlsx(mat, filename="total_portfolio_weights.xlsx", sheet_name="Portfolio")


def stock_prices_view(prices_df: pd.DataFrame):
    st.subheader("Stock Prices (from Column E)")
    st.dataframe(prices_df, use_container_width=True, hide_index=True)
    export_xlsx(prices_df, filename="stock_prices.xlsx", sheet_name="Prices")


# -------------------------------------------------
# Positions View (full version from your code)
# -------------------------------------------------
def positions_view(data, prices_df=None, groups_file=None):
    st.subheader("Positions View (Per-Group Sheets + Summary + Formulas)")

    if not data:
        st.info("No client data found. Please upload your main Excel file first.")
        return

    # ---------- 0) Groups mapping (optional) ----------
    def _norm_name(x):
        return str(x).strip().title()

    clients_df = pd.DataFrame({"Name": sorted(data.keys())})
    grouped_meta = None

    if groups_file is not None:
        try:
            if groups_file.name.lower().endswith(".csv"):
                gdf = pd.read_csv(groups_file)
            else:
                gdf = pd.read_excel(groups_file)

            colmap = {c.lower().strip(): c for c in gdf.columns}
            name_col = colmap.get("name")
            seq_col  = colmap.get("sequence")
            grp_col  = colmap.get("groups") or colmap.get("group")

            if not name_col:
                st.error("Groups file must contain a 'Name' column.")
                return

            gdf = gdf.rename(columns={
                name_col: "Name",
                (seq_col if seq_col else "Sequence"): "Sequence",
                (grp_col if grp_col else "Groups"): "Groups",
            })

            gdf["Name"] = gdf["Name"].map(_norm_name)
            if "Groups" not in gdf.columns:
                gdf["Groups"] = "Ungrouped"
            if "Sequence" not in gdf.columns:
                gdf["Sequence"] = None
            gdf["Sequence"] = pd.to_numeric(gdf["Sequence"], errors="coerce")

            clients_df["Name"] = clients_df["Name"].map(_norm_name)
            merged = clients_df.merge(
                gdf[["Name", "Groups", "Sequence"]],
                on="Name",
                how="left"
            )
            merged["Groups"]  = merged["Groups"].fillna("Ungrouped")
            merged["SeqSort"] = merged["Sequence"].fillna(10**9)
            grouped_meta = merged.sort_values(
                ["Groups", "SeqSort", "Name"]
            ).reset_index(drop=True)
        except Exception as e:
            st.warning(f"Could not read groups file, using default grouping. Details: {e}")

    if grouped_meta is None or grouped_meta.empty:
        grouped_meta = clients_df.copy()
        grouped_meta["Groups"]  = "Ungrouped"
        grouped_meta["SeqSort"] = 0
        grouped_meta["Sequence"] = None

    # ---------- 1) Price map from prices_df (optional) ----------
    price_map = {}
    if isinstance(prices_df, pd.DataFrame) and not prices_df.empty:
        if "Stock" in prices_df.columns and "Price" in prices_df.columns:
            price_map = dict(zip(prices_df["Stock"], prices_df["Price"]))

    def sanitize_sheet_name(s: str) -> str:
        s = re.sub(r'[:\\/?*\[\]]', "-", str(s)).strip()
        return (s or "Group")[:31]

    # Colors from Example.xlsx
    NAME_BAR_FILL      = XLFill(start_color="FF95A8C3", end_color="FF95A8C3", fill_type="solid")
    NAV_GREY_FILL      = XLFill(start_color="FFF4F4F4", end_color="FFF4F4F4", fill_type="solid")
    CASH_WHITE_FILL    = XLFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    STOCKS_HEADER_FILL = XLFill(start_color="FFCCE2F1", end_color="FFCCE2F1", fill_type="solid")
    STOCK_WHITE_FILL   = XLFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    STOCK_ALT_FILL     = XLFill(start_color="FFF4F4F4", end_color="FFF4F4F4", fill_type="solid")
    MOMENTUM_FILL      = XLFill(start_color="FFF4F4F4", end_color="FFF4F4F4", fill_type="solid")

    header = ["Item", "Value/Qty", "Weight", "MV"]
    all_preview_rows = []

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        created_any_sheet = False

        for grp_name, sub in grouped_meta.groupby("Groups", sort=False):
            sub_valid = sub[sub["Name"].isin(data.keys())].copy()
            if sub_valid.empty:
                sheet_name = sanitize_sheet_name(grp_name)
                base, idx = sheet_name, 1
                while sheet_name in writer.book.sheetnames:
                    sheet_name = sanitize_sheet_name(f"{base}-{idx}")
                    idx += 1
                pd.DataFrame([["Group", grp_name, None, None]], columns=header)\
                    .to_excel(writer, index=False, sheet_name=sheet_name)
                created_any_sheet = True
                continue

            rows = []
            block_markers = []

            group_total_cash = 0.0
            group_total_mv   = 0.0
            group_total_mom  = 0.0
            group_stocks_rows = []

            for _, rr in sub_valid.iterrows():
                client = rr["Name"]
                info   = data[client]

                cash_client = float(info.get("total_cash", 0) or info.get("cash", 0) or 0)
                mom_client  = float(info.get("momentum_mv", 0) or 0)
                dfc         = info["data"]

                total_mv_client = 0.0
                client_stock_rows = []
                if not dfc.empty:
                    for _, srow in dfc.iterrows():
                        stock = srow.get("Company Name")
                        qty   = float(srow.get("Quantity", 0) or 0)
                        row_price = float(srow.get("Price", 0) or 0)
                        row_mv    = float(srow.get("Market Value", 0) or 0)

                        eff_price = None
                        if stock in price_map and price_map[stock]:
                            eff_price = float(price_map[stock])
                        elif row_price:
                            eff_price = row_price

                        mv_calc = (qty * eff_price) if (eff_price not in [None, 0]) else row_mv
                        total_mv_client += mv_calc

                        client_stock_rows.append((stock, qty, eff_price, mv_calc))
                        group_stocks_rows.append({"Stock": stock, "Qty": qty, "Price": eff_price or 0.0, "MV": mv_calc})

                nav_client = cash_client + total_mv_client + mom_client

                group_total_cash += cash_client
                group_total_mv   += total_mv_client
                group_total_mom  += mom_client

                rows.append(["Group", grp_name, None, None])
                rows.append(["Name", client, None, None])
                nav_idx = len(rows)
                rows.append(["NAV", nav_client, None, None])
                cash_idx = len(rows)
                rows.append(["Total Cash", cash_client, None, None])
                rows.append(["Stocks", "Quantity", "Weight", "MV"])
                first_stock_idx = len(rows)

                if client_stock_rows:
                    for stock, qty, eff_price, mv_calc in client_stock_rows:
                        rows.append([stock, qty, None, None])
                    last_stock_idx = len(rows) - 1
                else:
                    last_stock_idx = None

                momentum_idx = len(rows)
                rows.append(["Momentum", mom_client, None, None])
                rows.append(["", "", "", ""])

                block_markers.append({
                    "nav_idx": nav_idx,
                    "cash_idx": cash_idx,
                    "first_stock_idx": first_stock_idx,
                    "last_stock_idx": last_stock_idx,
                    "momentum_idx": momentum_idx,
                })

            all_preview_rows.extend(rows + [["", "", "", ""]])

            grp_df = pd.DataFrame(rows, columns=header)

            sheet_name = sanitize_sheet_name(grp_name)
            base, idx = sheet_name, 1
            while sheet_name in writer.book.sheetnames:
                sheet_name = sanitize_sheet_name(f"{base}-{idx}")
                idx += 1

            grp_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=0, startcol=0)
            ws = writer.sheets[sheet_name]
            created_any_sheet = True

            # group summary
            group_total_nav = group_total_cash + group_total_mv + group_total_mom

            sum_df = pd.DataFrame(group_stocks_rows)
            if not sum_df.empty:
                gsum = sum_df.groupby("Stock", as_index=False).agg(
                    **{"Sum Qty": ("Qty", "sum"), "Sum MV": ("MV", "sum")}
                )
                gsum["Weight"] = gsum["Sum MV"].div(group_total_nav).fillna(0.0)
                if price_map:
                    gsum["Price"] = gsum["Stock"].map(price_map)
                    gsum["Price"] = gsum.apply(
                        lambda row: row["Price"] if row["Price"] not in [None, 0]
                        else (row["Sum MV"] / row["Sum Qty"] if row["Sum Qty"] else None),
                        axis=1
                    )
                else:
                    gsum["Price"] = gsum.apply(lambda row: (row["Sum MV"] / row["Sum Qty"]) if row["Sum Qty"] else None, axis=1)
            else:
                gsum = pd.DataFrame(columns=["Stock", "Sum Qty", "Sum MV", "Weight", "Price"])

            last_data_row = ws.max_row
            summary_start = last_data_row + 2

            cash_row    = summary_start + 1
            nav_row     = summary_start + 2
            mv_row      = summary_start + 3
            cash_mv_row = summary_start + 4

            ws.cell(row=summary_start, column=8, value="Group Summary")

            ws.cell(row=cash_row, column=8, value="Total Cash")
            cash_val_cell = ws.cell(row=cash_row, column=9, value=group_total_cash)
            cash_val_cell.number_format = "#,##0.00"

            ws.cell(row=nav_row, column=8, value="Total NAV")
            nav_val_cell = ws.cell(row=nav_row, column=9, value=group_total_nav)
            nav_val_cell.number_format = "#,##0.00"

            ws.cell(row=mv_row, column=8, value="Total MV")
            mv_val_cell = ws.cell(row=mv_row, column=9, value=group_total_mv)
            mv_val_cell.number_format = "#,##0.00"

            ws.cell(row=cash_row, column=10, value="Cash / NAV %")
            cash_nav_pct = ws.cell(row=cash_row, column=11)
            cash_nav_pct.value = f"=IFERROR({cash_val_cell.coordinate}/{nav_val_cell.coordinate},0)"
            cash_nav_pct.number_format = "0.00%"

            ws.cell(row=mv_row, column=10, value="MV / NAV %")
            mv_nav_pct = ws.cell(row=mv_row, column=11)
            mv_nav_pct.value = f"=IFERROR({mv_val_cell.coordinate}/{nav_val_cell.coordinate},0)"
            mv_nav_pct.number_format = "0.00%"

            ws.cell(row=cash_mv_row, column=8, value="Cash / MV %")
            cash_mv_pct = ws.cell(row=cash_mv_row, column=9)
            cash_mv_pct.value = f"=IFERROR({cash_val_cell.coordinate}/{mv_val_cell.coordinate},0)"
            cash_mv_pct.number_format = "0.00%"

            sum_header_row = summary_start + 8
            headers_sum = ["Stock", "Sum Qty", "Sum MV", "Weight", "Price"]
            for j, h in enumerate(headers_sum, start=8):
                ws.cell(row=sum_header_row, column=j, value=h)

            for i, row_sum in gsum.iterrows():
                rr2 = sum_header_row + 1 + i
                ws.cell(row=rr2, column=8, value=row_sum["Stock"])
                ws.cell(row=rr2, column=9, value=row_sum["Sum Qty"]).number_format = "#,##0"
                ws.cell(row=rr2, column=10, value=row_sum["Sum MV"]).number_format = "#,##0.00"
                ws.cell(row=rr2, column=11, value=row_sum["Weight"]).number_format = "0.00%"
                ws.cell(row=rr2, column=12, value=row_sum["Price"]).number_format = "#,##0.00"

            # summary styling
            for c in range(8, 13):
                ws.cell(row=summary_start, column=c).fill = STOCKS_HEADER_FILL
            metric_rows = [cash_row, nav_row, mv_row, cash_mv_row]
            for idx_m, r in enumerate(metric_rows):
                fill = STOCK_WHITE_FILL if idx_m % 2 == 0 else STOCK_ALT_FILL
                for c in range(8, 13):
                    ws.cell(row=r, column=c).fill = fill
            for c in range(8, 13):
                ws.cell(row=sum_header_row, column=c).fill = STOCKS_HEADER_FILL
            if not gsum.empty:
                alt = True
                for i in range(len(gsum)):
                    rr2 = sum_header_row + 1 + i
                    fill = STOCK_WHITE_FILL if alt else STOCK_ALT_FILL
                    for c in range(8, 13):
                        ws.cell(row=rr2, column=c).fill = fill
                    alt = not alt

            # price range for formulas
            if not gsum.empty:
                price_start = sum_header_row + 1
                price_end   = sum_header_row + len(gsum)
                price_range = f"$H${price_start}:$L${price_end}"
            else:
                price_range = None

            # formats
            for cidx in range(2, 5):
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=cidx)
                    if isinstance(cell.value, (int, float)):
                        if cidx == 3:
                            cell.number_format = "0.00%"
                        else:
                            cell.number_format = "#,##0" if float(cell.value).is_integer() else "#,##0.00"

            # formulas for NAV/MV/WT
            if price_range:
                for bm in block_markers:
                    nav_row_excel = bm["nav_idx"] + 2
                    cash_row_excel = bm["cash_idx"] + 2
                    momentum_row_excel = bm["momentum_idx"] + 2

                    nav_cell = ws.cell(row=nav_row_excel, column=2)
                    cash_cell = ws.cell(row=cash_row_excel, column=2)
                    momentum_cell = ws.cell(row=momentum_row_excel, column=2)

                    if bm["last_stock_idx"] is not None:
                        first_stock_row = bm["first_stock_idx"] + 2
                        last_stock_row = bm["last_stock_idx"] + 2
                        sum_mv_expr = f"SUM(D{first_stock_row}:D{last_stock_row})"
                        nav_cell.value = f"={cash_cell.coordinate}+{sum_mv_expr}+{momentum_cell.coordinate}"
                    else:
                        nav_cell.value = f"={cash_cell.coordinate}+{momentum_cell.coordinate}"
                    nav_cell.number_format = "#,##0.00"

                    if bm["last_stock_idx"] is not None:
                        for df_i in range(bm["first_stock_idx"], bm["last_stock_idx"] + 1):
                            excel_row = df_i + 2
                            stock_cell = ws.cell(row=excel_row, column=1)
                            qty_cell   = ws.cell(row=excel_row, column=2)
                            wt_cell    = ws.cell(row=excel_row, column=3)
                            mv_cell    = ws.cell(row=excel_row, column=4)

                            mv_cell.value = (
                                f'=IFERROR({qty_cell.coordinate}*'
                                f'VLOOKUP({stock_cell.coordinate},{price_range},5,FALSE),0)'
                            )
                            mv_cell.number_format = "#,##0.00"
                            wt_cell.value = f'=IFERROR({mv_cell.coordinate}/{nav_cell.coordinate},0)'
                            wt_cell.number_format = "0.00%"

            # main styling (A:D)
            for c in range(1, 5):
                ws.cell(row=1, column=c).fill = NAME_BAR_FILL

            in_stock_block = False
            alt_white = True
            for r in range(2, ws.max_row + 1):
                label = (ws.cell(row=r, column=1).value or "").strip().lower()

                if label in ("group", "name"):
                    for c in range(1, 5):
                        ws.cell(row=r, column=c).fill = NAME_BAR_FILL
                    in_stock_block = False
                elif label == "nav":
                    for c in range(1, 5):
                        ws.cell(row=r, column=c).fill = NAV_GREY_FILL
                    in_stock_block = False
                elif label == "total cash":
                    for c in range(1, 5):
                        ws.cell(row=r, column=c).fill = CASH_WHITE_FILL
                    in_stock_block = False
                elif label == "stocks":
                    for c in range(1, 5):
                        ws.cell(row=r, column=c).fill = STOCKS_HEADER_FILL
                    in_stock_block = True
                    alt_white = True
                elif label == "momentum":
                    for c in range(1, 5):
                        ws.cell(row=r, column=c).fill = MOMENTUM_FILL
                    in_stock_block = False
                elif label == "":
                    in_stock_block = False
                    continue
                else:
                    if in_stock_block:
                        fill = STOCK_WHITE_FILL if alt_white else STOCK_ALT_FILL
                        for c in range(1, 5):
                            ws.cell(row=r, column=c).fill = fill
                        alt_white = not alt_white

        if not created_any_sheet:
            pd.DataFrame(
                [["No data", "No matching clients in groups file"]],
                columns=["Item", "Value"],
            ).to_excel(writer, index=False, sheet_name="Summary")

    buf.seek(0)
    st.download_button(
        "📥 Download Excel (Positions by Group)",
        buf,
        file_name="positions_by_group.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    if all_preview_rows:
        preview_df = pd.DataFrame(all_preview_rows, columns=header)
        st.caption("Preview (structure matches Excel sheets):")
        st.dataframe(preview_df, hide_index=True, use_container_width=True)
    else:
        st.info("No rows to preview.")


# -------------------------------------------------
# Main App
# -------------------------------------------------
uploaded = st.file_uploader("Upload main Excel (.xlsx)", type=["xlsx"])

groups_file = st.file_uploader(
    "Optional: Upload Groups mapping (Sequence, Groups, Name)",
    type=["xlsx", "csv"],
    key="groups_mapping_top"
)

cashflow_file = st.file_uploader(
    "Optional: Upload Cashflow file (.xls)",
    type=["xls"],
    key="cashflow_xls"
)

cashflows_by_client = {}
if cashflow_file is not None:
    try:
        cashflows_by_client = load_cashflows_xls(cashflow_file)
        st.success("Cashflow file loaded.")
    except Exception as e:
        st.error(
            "Could not read the .xls cashflow file. "
            "This usually means xlrd is missing. Try: pip install xlrd\n\n"
            f"Details: {e}"
        )
        cashflows_by_client = {}

if uploaded:
    data, prices_df = extract_client_data(uploaded)

    view = st.selectbox(
        "Select View",
        [
            "Client View",
            "Total Portfolio View",
            "Total Portfolio View (Weights)",
            "Stock Prices View",
            "Positions View",
        ]
    )

    if view == "Client View":
        client_view(data, cashflows_by_client=cashflows_by_client)
    elif view == "Total Portfolio View":
        total_portfolio_view(data)
    elif view == "Total Portfolio View (Weights)":
        total_portfolio_view_weights(data)
    elif view == "Stock Prices View":
        stock_prices_view(prices_df)
    elif view == "Positions View":
        positions_view(data, prices_df, groups_file)
else:
    st.info("Please upload the main Excel file to begin.")






